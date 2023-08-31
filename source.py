import os # Used to make a directory and a file :)
from openpyxl import Workbook, load_workbook # Used to manage values in excel sheet # https://openpyxl.readthedocs.io/en/stable/index.html
path = os.path.join(os.path.expanduser('~') + "/Documents/","Savings-Docs") #stores values locally for security
file_path = path + "/savings-log.xlsx"

# Create the file directory if it does not exist
# and create the file with user provided values
def create_new():
    
    try:
        os.mkdir(path)
        print("Creating new directory...")
    except OSError as error:
        print("Accessing existing directory...")

    if not os.path.exists(file_path):
        print("Creating new file...")
        wb = Workbook()
        ws = wb.active
        ws['B1'] = "Amount ($)"
        ws['C1'] = "Locked?"
        ws['A2'] = "Total Balance"
        ws['B2'] = float(input("What is your total balance? $"))
        ws['C2'] = "false"
        ws['A3'] = "Unaccounted Funds"
        ws['B3'] = ws['B2'].value
        ws['C3'] = "false"
        ws['A4'] = "Num of Categories"
        ws['B4'] = int(input("How many savings categories would you like to have? "))
        ws['C4'] = "false"
        
        for i in range(1,ws['B4'].value + 1):
            if i%10 == 1:
                cat = str(input("What would you like to name your " + str(i) + "st category? "))
            elif i%10 == 2:
                cat = str(input("What would you like to name your " + str(i) + "nd category? "))
            elif i%10 == 3:
                cat = str(input("What would you like to name your " + str(i) + "rd category? "))
            else:
                cat = str(input("What would you like to name your " + str(i) + "th category? "))

            locked = str(input("Is this cell locked? y/n (Exlcuded from fund redistribution feature) "))
            if locked == 'y':
                locked = "true"
            else:
                locked = "false"
            
            ws['A' + str(5 + i)] = cat
            ws['B' + str(5 + i)] = 0
            ws['C' + str(5 + i)] = locked
        
        wb.save(file_path) 
        print("File creation successful...")  
        
def allocate_funds(ws):
    if ws['B3'].value > 0:
        if input("You have an unaccounted for balance of $" + str(ws['B3'].value) + ", would you like to desigate a category? y/n ") == 'y':
            print("Which categories would you like to depoist into?")
            for i in range(ws['B4'].value): # str(ws['A' + str(i + 6)].value)
                print(str(i+1) + ". " + str(ws['A' + str(i + 6)].value))
            ans = list(map(int, input("Please select the category(ies) above that you wish to deposit into (i.e. \"1 2 3\"): ").strip().split()))[:ws['B4'].value]
            for i in ans:
                k = i -1
                check = True
                while check:
                    amt = float(input("How much would you like to put in the " + str(ws['A' + str((i-1) + 6)].value) + " category? $"))
                    if amt == ws['B3'].value:
                        print("Funds have successfully been alocated.")
                        ws['B' + str(k + 6)] = amt + ws['B' + str(k + 6)].value
                        ws['B3'] = ws['B3'].value - amt
                        check = False
                        break
                    elif amt < ws['B3'].value:
                        print("Funds have successfully been allocated but there is remaining money left unallocated.")
                        ws['B' + str(k + 6)] = amt + ws['B' + str(k + 6)].value
                        ws['B3'] = ws['B3'].value - amt
                        check = False
                    elif amt > ws['B3'].value:
                        print("You cannot allocate more funds than you have deposited. Please try again.")
                        
def withdraw_category(ws, debt):
    while debt > 0:
            print("Which category would you like to withdraw from? (Numbered)")
            for i in range(1, ws['B4'].value + 1):
                print(str(i) + ". " + ws['A' + str(i+5)].value + ": $" +  str(ws['B' + str(i+5)].value))
            pull = int(input())
            amt = float(input("How much would you like to withdraw from " + ws['A' + str(pull+5)].value + "? $"))
            debt -= amt
            if amt <= ws['B' + str(pull+5)].value:
                print("Withdrawal from " + ws['A' + str(pull+5)].value + " was successful.")
            else:
                print("You have withdrawn $" + str(amt) + " from " + ws['A' + str(pull+5)].value + ". Your balance in that category did not cover the withdraw and is now negative.")
                print("It is reccomended you move money from other categories to cover this deficit or to allocate money during your next deposit.")
            ws['B' + str(pull+5)] = ws['B' + str(pull+5)].value - amt

def deposit(ws):
    if input("Have you made a deposit into savings that you would like to record? y/n ") == 'y':
        amt = float(input("How much? $"))
        ws['B2'] = ws['B2'].value + amt
        ws['B3'] = ws['B3'].value + amt
        print("Your new total balance is $" + str(ws['B2'].value))

def withdrawl(ws):
    if input("Have you made a withdrawal from savings that you would like to record? y/n ") == 'y':
        debt = float(input("How much? $"))
        ws['B2'] = ws['B2'].value - debt
        if ws['B3'].value >= debt:
            ws['B3'] = ws['B3'].value - debt
            print("The money withdrawal has been removed from the unaccounted for funds, no further action is needed.")
        elif ws['B3'].value > 0:
            debt -= ws['B3'].value
            ws['B3'] =  0
            print("Some of the money withdrawal has been removed from the unaccounted for funds, further action is needed for the remaining $" + str(debt) + " withdrawal balance.")
            withdraw_category(ws, debt)
        else:
            withdraw_category(ws, debt)